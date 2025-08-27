import logging

import openpyxl

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def test_field_replacement():
    """Test field replacement logic on sample file"""

    # Load sample file
    file_path = 'sample_data/reports_excel/report_34300798_2025.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    print(f"Testing field replacement on: {file_path}")
    print(f"Sheet size: {ws.max_row} rows, {ws.max_column} columns")
    print()

    # Show current content
    print("CURRENT CONTENT:")
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_val = cell.value if cell.value is not None else 'None'
            row_values.append(f'[{col_idx}]:"{cell_val}"')
        print(f'Row {row_idx}: {" | ".join(row_values)}')
    print()

    # Test data for replacement
    replacement_data = {
        'אסמכתת בסיס': '226239',
        "סה'כ אריזות": '0.2',
        "סה'כ משקל": '1',
        'שם כרטיס': 'Test Name',
        'שם לועזי': 'Test Foreign Name',
        'כתובת': 'Test Address',
    }

    # Define field replacements
    field_replacements = [
        {
            'search_fields': ['אסמכתת בסיס', 'אסמכתת בסיס'],
            'replace_value': str(replacement_data.get('אסמכתת בסיס', 'אסמכתת בסיס'))
        },
        {
            'search_fields': ['סה\'כ משקל', 'סהכ משקל', 'סה"כ משקל', 'סהכ משקל', 'משקל כולל', 'סהכ משקל'],
            'replace_value': str(replacement_data.get('סה\'כ משקל', replacement_data.get('סהכ משקל', 'סה\'כ משקל')))
        },
        {
            'search_fields': ['סה\'כ אריזות', 'סהכ אריזות', 'סה"כ אריזות', 'סהכ אריזות', 'כמות אריזות', 'סהכ אריזות'],
            'replace_value': str(replacement_data.get('סה\'כ אריזות', replacement_data.get('סהכ אריזות', 'סה\'כ אריזות')))
        }
    ]

    print("FIELD REPLACEMENTS CONFIG:")
    for replacement in field_replacements:
        print(f"  Fields: {replacement['search_fields']}")
        print(f"  Value: '{replacement['replace_value']}'")
    print()

    # Test the first part of the logic (first two columns search)
    print("TESTING FIRST PART (first two columns search):")
    replacements_made = 0

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        cells = list(row)
        if len(cells) >= 2:  # Need at least 2 columns
            field_cell = cells[0]  # Column "שדה"
            value_cell = cells[1]  # Column "ערך"

            if field_cell.value is not None:
                field_value = str(field_cell.value)
                print(f"Row {row_idx}: field='{field_value}', value='{value_cell.value}'")

                # Check if this field matches any of our search fields
                for replacement in field_replacements:
                    for search_field in replacement['search_fields']:
                        if field_value == search_field:
                            # Replace the value in the adjacent cell
                            old_value = value_cell.value
                            value_cell.value = replacement['replace_value']
                            replacements_made += 1
                            print(f"  ✓ REPLACED: '{search_field}' from '{old_value}' to '{replacement['replace_value']}'")
                            break
                    else:
                        continue
                    break

    print(f"Replacements made in first part: {replacements_made}")
    print()

    # Show content after first part
    print("CONTENT AFTER FIRST PART:")
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_val = cell.value if cell.value is not None else 'None'
            row_values.append(f'[{col_idx}]:"{cell_val}"')
        print(f'Row {row_idx}: {" | ".join(row_values)}')
    print()

    # Save test results
    test_output_path = 'test_replacement_result.xlsx'
    wb.save(test_output_path)
    wb.close()

    print(f"Test results saved to: {test_output_path}")

if __name__ == "__main__":
    test_field_replacement()

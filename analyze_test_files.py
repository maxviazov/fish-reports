#!/usr/bin/env python3
"""
Analyze the test files to understand why replacement is not working.
"""

import openpyxl
import pandas as pd


def main():
    # Сначала посмотрим на filtered_data.xlsx
    print('=== Данные из filtered_data.xlsx ===')
    df = pd.read_excel('test_output/filtered_data.xlsx')
    print(df.head())

    # Найдем данные для ולדמן אשדוד
    license_data = df[df['ח"פ לקוח או מספר עוסק'] == 512642182]
    if not license_data.empty:
        print('\n=== Данные для ולדמן אשדוד (license: 512642182) ===')
        for _, row in license_data.iterrows():
            print('אסמכתת בסיס:', row['אסמכתת בסיס'])
            print('סהכ משקל:', row['סהכ משקל'])
            print('סהכ אריזות:', row['סהכ אריזות'])
    else:
        print('Данные для license 512642182 не найдены')

    # Теперь посмотрим на структуру файла ולדמן אשדוד.xlsx
    print('\n=== Структура файла ולדמן אשדוד.xlsx ===')
    wb = openpyxl.load_workbook('test_output/ולדמן אשדוד.xlsx')
    ws = wb.active

    print(f'Лист: {ws.title}')
    print(f'Размер: {ws.max_row} строк, {ws.max_column} колонок')

    # Покажем все непустые ячейки
    print('\nВсе непустые ячейки:')
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                print(f'[{row},{col}]: "{cell_value}"')

    wb.close()

if __name__ == "__main__":
    main()

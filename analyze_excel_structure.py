import openpyxl

# Load one of the sample Excel files
file_path = 'sample_data/reports_excel/report_512451451_2025.xlsx'
try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    print(f'Анализ файла: {file_path}')
    print(f'Размер листа: {ws.max_row} строк, {ws.max_column} колонок')
    print()

    # Show all rows
    print('ВСЕ СТРОКИ:')
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_val = cell.value if cell.value is not None else 'None'
            row_values.append(f'[{col_idx}]:"{cell_val}"')
        print(f'Строка {row_idx}: {" | ".join(row_values)}')

    wb.close()

except Exception as e:
    print(f'Ошибка: {e}')

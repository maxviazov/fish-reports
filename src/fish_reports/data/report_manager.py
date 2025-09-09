"""
Report management utilities for Fish Reports processing.
"""

import logging
import re
import shutil

# from datetime import datetime  # Импорт внутри функции для избежания предупреждений линтера
from pathlib import Path
from typing import Dict, List, Optional, Union

import openpyxl
import pandas as pd

from ..utils import CityManager

logger = logging.getLogger(__name__)


class ReportManager:
    """Manages report files and data replacement operations."""

    def __init__(self, base_dir: Path, output_dir: Path, city_manager=None):
        """
        Initialize the report manager.

        Args:
            base_dir: Base directory for the project
            output_dir: Directory for output files
            city_manager: CityManager instance for city validation
        """
        self.base_dir = Path(base_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.copied_files_count = 0

        # Initialize city manager for city validation
        if city_manager:
            self.city_manager = city_manager
        else:
            self.city_manager = CityManager()

    def process_reports(self, reports_dir: Path, intermediate_file: Path) -> Dict[str, str]:
        """
        Process all report files by replacing fields with data from intermediate file.
        Now supports multiple addresses per license by creating separate files.

        Args:
            reports_dir: Directory containing report template files
            intermediate_file: Path to Excel file with replacement data

        Returns:
            Dictionary mapping source file names to result file paths
        """
        results = {}

        if not reports_dir.exists():
            logger.error(f"Reports directory not found: {reports_dir}")
            return results

        if not intermediate_file.exists():
            logger.error(f"Intermediate file not found: {intermediate_file}")
            return results

        # Load replacement data
        logger.info(f"Loading replacement data from: {intermediate_file}")
        license_data_map = self._load_intermediate_data(intermediate_file)

        if not license_data_map:
            logger.error("No valid replacement data found")
            return results

        logger.info(f"Loaded data for {len(license_data_map)} unique license-address combinations")

        # Group data by license number
        license_groups = {}
        for unique_key, data in license_data_map.items():
            license_num = data['license_num']
            if license_num not in license_groups:
                license_groups[license_num] = []
            license_groups[license_num].append((unique_key, data))

        logger.info(f"Grouped into {len(license_groups)} license groups")

        # Process each license group
        for license_num, data_list in license_groups.items():
            logger.info(f"Processing license {license_num} with {len(data_list)} address combinations")

            # Find report files for this license
            license_files = self._find_files_for_license(reports_dir, license_num)

            if not license_files:
                logger.warning(f"No report files found for license: {license_num}")
                continue

            # Process each address combination
            for idx, (unique_key, replacement_data) in enumerate(data_list):
                address_key = replacement_data.get('address_key', 'no_address')
                logger.info(f"Processing address combination {idx+1}/{len(data_list)}: {address_key}")

                # Extract city code from address for validation
                expected_city_code = self._extract_city_code_from_address(replacement_data.get('כתובת', ''))

                # Filter files by city if we have multiple files and a city code
                if len(license_files) > 1 and expected_city_code:
                    logger.info(f"Multiple files found for license {license_num}, filtering by city code: {expected_city_code}")
                    filtered_files = self._filter_files_by_city(license_files, expected_city_code)

                    if not filtered_files:
                        logger.warning(f"No files match city code {expected_city_code} for license {license_num}, using all files")
                        filtered_files = license_files
                else:
                    filtered_files = license_files

                logger.info(f"Using {len(filtered_files)} file(s) for license {license_num}, address: {address_key}")

                # For each file and each address, create a separate output file
                for file_path in filtered_files:
                    try:
                        # Create unique output filename
                        base_name = file_path.stem
                        if len(data_list) > 1:
                            # Multiple addresses - include address in filename
                            if address_key != 'no_address':
                                new_name = f"{base_name}_{address_key}.xlsx"
                            else:
                                new_name = f"{base_name}_{idx+1}.xlsx"
                        else:
                            # Single address - keep original name
                            new_name = f"{base_name}.xlsx"

                        dest_path = self.output_dir / new_name

                        # Perform replacement
                        logger.info(f"Processing: {file_path.name} -> {new_name} (license: {license_num}, address: {address_key})")
                        success = self._copy_file_with_replacement(file_path, dest_path, replacement_data)

                        if success:
                            results[file_path.name] = str(dest_path)
                            logger.info(f"Successfully processed: {new_name}")
                            self.copied_files_count += 1
                        else:
                            logger.error(f"Failed to process: {new_name}")

                    except Exception as e:
                        logger.error(f"Error processing {file_path.name}: {e}")

        logger.info(f"Processing complete. Successfully processed {len(results)} files")

        # Report unprocessed licenses if any
        self._report_unprocessed_licenses_new(license_data_map, results)

        return results

    def _extract_city_from_report(self, file_path: Path) -> Optional[str]:
        """
        Extract city code from a report file.

        Args:
            file_path: Path to the report file

        Returns:
            City code as string, or None if not found
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            # Look for city code patterns (typically 2-3 digits)
                            city_match = re.search(r'\b(\d{2,3})\b', cell)
                            if city_match:
                                city_code = city_match.group(1)
                                # Validate that this is actually a city code
                                if self.city_manager.is_valid_city_code(city_code):
                                    wb.close()
                                    return city_code
            wb.close()
        except Exception as e:
            logger.debug(f"Error extracting city from {file_path}: {e}")

        return None

    def _filter_files_by_city(self, file_paths: List[Path], expected_city_code: str) -> List[Path]:
        """
        Filter report files by city code to ensure correct city matching.

        Args:
            file_paths: List of file paths to filter
            expected_city_code: Expected city code from address data

        Returns:
            Filtered list of file paths that match the expected city
        """
        if not expected_city_code:
            logger.info("No city code provided for filtering, returning all files")
            return file_paths

        filtered_files = []
        logger.info(f"Filtering {len(file_paths)} files for city code: {expected_city_code}")

        for file_path in file_paths:
            try:
                report_city_code = self._extract_city_from_report(file_path)
                if report_city_code:
                    if report_city_code == expected_city_code:
                        filtered_files.append(file_path)
                        logger.info(f"File {file_path.name} matches city code {expected_city_code}")
                    else:
                        logger.info(f"File {file_path.name} has city code {report_city_code}, expected {expected_city_code} - excluded")
                else:
                    logger.warning(f"Could not extract city code from {file_path.name}, including in results")
                    filtered_files.append(file_path)
            except Exception as e:
                logger.error(f"Error processing {file_path.name} for city validation: {e}")
                # Include file if we can't validate it
                filtered_files.append(file_path)

        logger.info(f"City filtering complete: {len(filtered_files)}/{len(file_paths)} files match")
        return filtered_files

    def _extract_city_code_from_address(self, full_address: str) -> Optional[str]:
        """
        Extract city code from a full address string.

        Args:
            full_address: Full address string (e.g., "אשדוד, העצמאות 87")

        Returns:
            City code as string, or None if not found
        """
        if not full_address or not isinstance(full_address, str):
            return None

        # Try to extract city name from address (before comma)
        if ',' in full_address:
            city_part = full_address.split(',')[0].strip()
            logger.info(f"Extracted city name from address: '{city_part}'")

            # Get city code from city manager
            city_code = self.city_manager.get_city_code_by_name(city_part)
            if city_code:
                logger.info(f"Found city code {city_code} for city '{city_part}'")
                return city_code
            else:
                logger.debug(f"City '{city_part}' not found in city database")
        else:
            logger.debug(f"No comma found in address '{full_address}', cannot extract city")

        return None

    def _find_files_for_license(self, reports_dir: Path, license_num: str) -> List[Path]:
        """
        Find all report files that contain the given license number.
        Now supports city validation when multiple files are found.

        Args:
            reports_dir: Directory to search for report files
            license_num: License number to search for

        Returns:
            List of file paths containing the license
        """
        matching_files = []

        # Search by filename
        for pattern in ("*.xlsx", "*.xlsm"):
            for file_path in reports_dir.glob(pattern):
                if license_num in file_path.name:
                    matching_files.append(file_path)

        # If no files found by filename, search by content
        if not matching_files:
            logger.info(f"No files found by filename for license {license_num}, searching by content...")
            try:
                for pattern in ("*.xlsx", "*.xlsm"):
                    for file_path in reports_dir.glob(pattern):
                        try:
                            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                            found = False
                            for ws in wb.worksheets:
                                for row in ws.iter_rows(values_only=True):
                                    for cell in row:
                                        if cell and license_num in str(cell):
                                            matching_files.append(file_path)
                                            found = True
                                            break
                                    if found:
                                        break
                            wb.close()
                            if found:
                                break
                        except Exception:
                            continue
            except Exception as e:
                logger.error(f"Error searching by content: {e}")

        return matching_files

    def _report_unprocessed_licenses_new(self, license_data_map: Dict[str, Dict], results: Dict[str, str]):
        """
        Report licenses that have data but no corresponding report files.
        Updated to work with new license-address key structure.

        Args:
            license_data_map: Dictionary with all available license data
            results: Dictionary with successfully processed files
        """
        # Extract license numbers from processed files
        processed_licenses = set()
        for filename in results.keys():
            license_num = self._extract_license_from_filename(filename)
            if license_num:
                processed_licenses.add(license_num)

        # Get all unique licenses from data
        all_licenses = set()
        for unique_key, data in license_data_map.items():
            all_licenses.add(data['license_num'])

        # Find unprocessed licenses
        unprocessed_licenses = all_licenses - processed_licenses

        if unprocessed_licenses:
            logger.warning("Found licenses with data but no corresponding report files:")
            for license_num in sorted(unprocessed_licenses):
                # Find all address combinations for this license
                license_entries = [(k, v) for k, v in license_data_map.items() if v['license_num'] == license_num]
                logger.warning(f"  License {license_num}: {len(license_entries)} address combinations")
                for unique_key, data in license_entries:
                    address = data.get('כתובת', 'N/A')
                    logger.warning(f"    Address: {address}")
                    logger.warning(f"    Base document: {data.get('אסמכתת בסיס', 'N/A')}")
                    logger.warning("    Total packages: %s" % data.get("סה'כ אריזות", 'N/A'))
                    logger.warning("    Total weight: %s" % data.get("סה'כ משקל", 'N/A'))
        else:
            logger.info("All licenses with data have corresponding report files")

    def _extract_license_from_filename(self, filename: str) -> Optional[str]:
        """
        Extract license number from filename.

        Args:
            filename: Name of the report file

        Returns:
            License number as string, or None if not found
        """
        # Try different patterns to extract license number
        patterns = [
            r'(\d{9})',  # 9-digit number
            r'(\d{8})',  # 8-digit number
            r'(\d{7})',  # 7-digit number
        ]

        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                return match.group(1)

        return None

    def _format_address_for_ministry(self, full_address: str) -> str:
        """
        Format address for Ministry of Health requirements.
        Removes city name and keeps only street address.

        Args:
            full_address: Full address string (e.g., "אשדוד, העצמאות 87")

        Returns:
            Street address only (e.g., "העצמאות 87")
        """
        if not full_address or not isinstance(full_address, str):
            return ""

        # Split by comma and take the part after comma (street address)
        parts = full_address.split(',', 1)
        if len(parts) > 1:
            # Take the street part and strip whitespace
            street_address = parts[1].strip()
            logger.info(f"Formatted address: '{full_address}' -> '{street_address}'")
            return street_address
        else:
            # No comma found, return as is (might already be street-only)
            logger.info(f"Address already street-only: '{full_address}'")
            return full_address.strip()

    def _load_intermediate_data(self, intermediate_file: Path) -> Dict[str, Dict]:
        """
        Load data from intermediate Excel file for field replacement.

        Args:
            intermediate_file: Path to Excel file with data

        Returns:
            Dictionary mapping license numbers to replacement data
        """
        try:
            df = pd.read_excel(intermediate_file)

            # Check required columns
            required_cols = ['מספר עוסק מורשה', 'אסמכתת בסיס']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                logger.error(f"Missing required columns in intermediate file: {missing_cols}")
                return {}

            # Create mapping: license -> data
            license_data = {}
            address_column = None

            # Find address column
            for col in df.columns:
                if 'כתובת' in str(col):
                    address_column = col
                    break

            for _, row in df.iterrows():
                # Convert license number properly - handle both int and float types
                license_raw = row['מספר עוסק מורשה']
                if pd.isna(license_raw):
                    continue

                # Convert to int first to remove any decimal points, then to string
                try:
                    license_num = str(int(float(license_raw)))
                except (ValueError, TypeError):
                    license_num = str(license_raw)

                # Get address as selected by user (no parsing needed)
                address = ""
                if address_column:
                    full_address = row[address_column] if address_column in row else ""
                    # Format address for Ministry requirements (remove city name)
                    address = self._format_address_for_ministry(full_address)

                # Create unique key for license + address combination
                if address:
                    # Use formatted address as key (without city name)
                    address_key = str(address).strip()
                    unique_key = f"{license_num}_{address_key}"
                else:
                    # Fallback for empty address
                    address_key = "no_address"
                    unique_key = f"{license_num}_{address_key}"

                # Create data entry
                data_entry = {
                    'אסמכתת בסיס': row['אסמכתת בסיס'],
                    'סה\'כ אריזות': row.get('סה\'כ אריזות', 0),
                    'סה\'כ משקל': row.get('סה\'כ משקל', 0),
                    'שם כרטיס': row.get('שם כרטיס', ''),
                    'שם לועזי': row.get('שם לועזי', ''),
                    'כתובת': address,
                    'license_num': license_num,
                    'address_key': address_key
                }

                # Store with unique key
                license_data[unique_key] = data_entry

                logger.info(f"Loaded data for license {license_num}, address: {address_key}")

            logger.info(f"Total unique license-address combinations: {len(license_data)}")
            return license_data

        except Exception as e:
            logger.error(f"Error loading intermediate data: {e}")
            return {}

    def _copy_file_with_replacement(self, source_path: Path, dest_path: Path, replacement_data: Dict) -> bool:
        """
        Copy file and replace data during the copy process.

        Args:
            source_path: Source Excel file path
            dest_path: Destination Excel file path
            replacement_data: Dictionary with replacement values

        Returns:
            True if successful, False otherwise
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(source_path)

            # Get field mappings for this file type
            field_mappings = self._get_field_mappings()

            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self._replace_fields_in_worksheet(worksheet, replacement_data, field_mappings)

            # Save to destination
            workbook.save(dest_path)
            workbook.close()

            # Verify the file was saved correctly
            logger.info(f"Файл сохранен: {dest_path}")
            try:
                # Quick verification - try to read back the file
                verify_wb = openpyxl.load_workbook(dest_path, data_only=True)
                verify_sheet = verify_wb.active

                # Проверяем значения в ключевых числовых полях
                for row_idx, row in enumerate(verify_sheet.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value is not None:
                            cell_value = str(cell.value)
                            # Проверяем ключевые поля
                            if 'מוצרים מוכנים לאכילה' in cell_value or 'סה"כ קרטונים' in cell_value:
                                # Проверяем следующую ячейку в строке
                                if col_idx < len(row):
                                    next_cell = row[col_idx]
                                    if next_cell.value is not None:
                                        logger.info(f"Проверка поля '{cell_value}': значение = {next_cell.value} (тип: {type(next_cell.value)})")

                logger.info(f"Проверка файла: {verify_sheet.max_row} строк, {verify_sheet.max_column} колонок")
                verify_wb.close()
                logger.info("Файл успешно сохранен и проверен")
            except Exception as e:
                logger.error(f"Ошибка при проверке сохраненного файла: {e}")

            return True

        except Exception as e:
            logger.error(f"Error copying file with replacement: {e}")
            return False

    def _get_field_mappings(self) -> Dict[str, str]:
        """
        Get mapping of field names to replacement keys.

        Returns:
            Dictionary mapping field names to data keys
        """
        field_mappings = {
            # Base document reference - mapping from intermediate to final report
            'אסמכתת בסיס': 'אסמכתת בסיס',
            'מספר תעודת משלוח': 'אסמכתת בסיס',  # Final report field -> intermediate data

            # Package and weight totals - mapping from intermediate to final report
            'סה\'כ אריזות': 'סה\'כ אריזות',
            'סה"כ קרטונים': 'סה\'כ אריזות',  # Final report field -> intermediate data
            'סה\'כ משקל': 'סה\'כ משקל',
            'מוצרים מוכנים לאכילה': 'סה\'כ משקל',  # Final report field -> intermediate data

            # Client information
            'שם כרטיס': 'שם כרטיס',
            'שם לועזי': 'שם לועזי',
            'כתובת': 'כתובת',
        }

        return field_mappings

    def _get_weight_value(self, replacement_data: Dict) -> float:
        """
        Get weight value from replacement data with proper handling of missing/zero values.

        Args:
            replacement_data: Dictionary with replacement values

        Returns:
            Weight value as float, 0 if missing or invalid
        """
        # Try different variations of weight field name
        weight_keys = ['סה\'כ משקל', 'סהכ משקל', 'סה"כ משקל', 'סהכ משקל']

        for key in weight_keys:
            if key in replacement_data:
                value = replacement_data[key]
                try:
                    # Convert to float, handling various formats
                    if pd.isna(value) or value is None or str(value).strip() == '':
                        continue
                    numeric_value = float(value)
                    if numeric_value >= 0:  # Accept zero and positive values
                        logger.info(f"Найдено значение веса '{key}': {numeric_value}")
                        return numeric_value
                except (ValueError, TypeError):
                    logger.warning(f"Не удалось преобразовать значение веса '{value}' в число")
                    continue

        # If no valid weight found, return 0
        logger.info("Значение веса не найдено или равно 0, будет установлено значение 0")
        return 0.0

    def _replace_fields_in_worksheet(self, worksheet, replacement_data: Dict, field_mappings: Dict[str, str]):
        """
        Replace fields in a single worksheet.

        Args:
            worksheet: openpyxl worksheet object
            replacement_data: Dictionary with replacement values
            field_mappings: Dictionary mapping field names to data keys
        """
        # Define field replacements for specific known fields
        # Ищем несколько вариантов написания полей
        logger.info(f"Данные для замены: {replacement_data}")

        # Define field replacements for specific known fields with Hebrew character handling
        # Mapping from intermediate file fields to final report fields
        logger.info(f"Данные для замены: {replacement_data}")

        # Получаем текущую дату в формате dd.mm.yy (больше не используется, заменено на формулу)
        # current_date = datetime.now().strftime('%d.%m.%y')

        weight_value = self._get_weight_value(replacement_data)

        field_replacements = [
            # אסמכתת בסיס -> מספר תעודת משלוח
            {
                'intermediate_field': 'אסמכתת בסיס',
                'target_column': 'מספר תעודת משלוח',
                'search_fields': ['מספר תעודת משלוח', 'אסמכתת בסיס'],
                'replace_value': replacement_data.get('אסמכתת בסיס', ''),
                'is_numeric': False  # Это текстовое поле
            },
            # סה'כ משקל -> מוצרים מוכנים לאכילה
            {
                'intermediate_field': 'סה\'כ משקל',
                'target_column': 'מוצרים מוכנים לאכילה',
                'search_fields': ['מוצרים מוכנים לאכילה', 'משקל כולל', 'משקל'],
                'replace_value': weight_value,
                'is_numeric': True,
                'force_replace': True
            },
            # סה'כ משקל -> סה"כ משקל
            {
                'intermediate_field': 'סה\'כ משקל',
                'target_column': 'סה"כ משקל',
                'search_fields': ['סה"כ משקל', 'סה\'כ משקל', 'סהכ משקל'],
                'replace_value': weight_value,
                'is_numeric': True,
                'force_replace': True
            },
            # סה'כ אריזות -> סה"כ קרטונים
            {
                'intermediate_field': 'סה\'כ אריזות',
                'target_column': 'סה"כ קרטונים',
                'search_fields': ['סה"כ קרטונים', 'סה\'כ אריזות', 'סהכ אריזות', 'כמות אריזות'],
                'replace_value': replacement_data.get('סה\'כ אריזות', replacement_data.get('סהכ אריזות', 0)),
                'is_numeric': True  # Это числовое поле
            },
            # כתובת -> כתובת (адрес)
            {
                'intermediate_field': 'כתובת',
                'target_column': 'כתובת',
                'search_fields': ['כתובת', 'address', 'כתובת'],
                'replace_value': replacement_data.get('כתובת', ''),
                'is_numeric': False  # Это текстовое поле
            },
            # תאריך -> תאריך (текущая дата с правильным форматом)
            {
                'intermediate_field': 'current_date_formula',
                'target_column': 'תאריך',
                'search_fields': ['תאריך', 'date', 'תאריך'],
                'replace_value': 'PLACEHOLDER_DATE',  # Будет заменено на текущую дату
                'is_numeric': False,  # Это дата, не число
                'is_formula': True  # Флаг для специальной обработки даты
            }
        ]

        logger.info("Конфигурация замен по столбцам:")
        for replacement in field_replacements:
            value_display = replacement['replace_value']
            if replacement['is_numeric'] and value_display != '':
                try:
                    # Для числовых значений показываем как число
                    numeric_value = float(value_display) if value_display != '' else 0
                    value_display = f"{numeric_value} (число)"
                except (ValueError, TypeError):
                    value_display = f"{value_display} (число, ошибка преобразования)"
            logger.info(f"  {replacement['intermediate_field']} -> столбец '{replacement['target_column']}' : '{value_display}'")

        # Log which fields we have data for
        available_data = {k: v for k, v in replacement_data.items() if v is not None and str(v).strip()}
        logger.info(f"Доступные данные для замены: {available_data}")

        replacements_made = 0

        # Log worksheet info for debugging
        logger.info(f"Анализируем лист '{worksheet.title}' с {worksheet.max_row} строками и {worksheet.max_column} колонками")

        # Сначала найдем номера столбцов по их названиям
        column_mapping = {}
        header_row = None

        # Ищем строку заголовков (обычно первая строка)
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            row_cells = list(row)
            if len(row_cells) >= 3:  # Минимум 3 колонки
                # Проверяем, есть ли в строке названия наших целевых столбцов
                row_text = ' '.join([str(cell.value) if cell.value is not None else '' for cell in row_cells])
                has_target_columns = any(target_col in row_text for replacement in field_replacements for target_col in [replacement['target_column']])

                if has_target_columns:
                    header_row = row_idx
                    logger.info(f"Найдена строка заголовков: {row_idx}")

                    # Определяем номера столбцов
                    for col_idx, cell in enumerate(row_cells, 1):
                        if cell.value is not None:
                            cell_value = str(cell.value)
                            for replacement in field_replacements:
                                target_col = replacement['target_column']
                                # Handle Hebrew special characters
                                normalized_cell = cell_value.replace('"', "'").replace('״', "'")
                                normalized_target = target_col.replace('"', "'").replace('״', "'")

                                if target_col in cell_value or normalized_target in normalized_cell:
                                    column_mapping[target_col] = col_idx
                                    logger.info(f"Столбец '{target_col}' найден в колонке {col_idx}")
                    break

        if not column_mapping:
            logger.warning("Не найдены целевые столбцы в файле - используем альтернативную стратегию поиска")
            # Сначала попробуем добавить отсутствующие поля
            replacements_made += self._add_missing_fields(worksheet, field_replacements, replacement_data)
            # Затем используем альтернативную стратегию поиска
            replacements_made += self._search_fields_in_all_cells(worksheet, field_replacements, replacement_data)
            return replacements_made

        logger.info(f"Найденные столбцы: {column_mapping}")

        # Теперь заменяем значения в найденных столбцах
        data_row_idx = header_row + 1 if header_row else 2  # Предполагаем, что данные начинаются со следующей строки

        for replacement in field_replacements:
            target_col = replacement['target_column']
            if target_col in column_mapping:
                col_idx = column_mapping[target_col]

                # Ищем строку с данными (обычно вторая строка после заголовков)
                for row_offset in [0, 1, 2]:  # Проверяем несколько строк
                    try:
                        data_row_idx = header_row + 1 + row_offset
                        if data_row_idx > worksheet.max_row:
                            break

                        cell = worksheet.cell(row=data_row_idx, column=col_idx)
                        if cell.value is not None or replacement.get('force_replace', False):
                            old_value = cell.value if cell.value is not None else ''

                            # Специальная обработка для поля веса - всегда проверяем на "חסרים משקלים"
                            if target_col == 'מוצרים מוכנים לאכילה' and str(old_value).strip() in ['חסרים משקלים', 'חסרים משקלים']:
                                    # Принудительно заменяем "חסרים משקלים" на числовое значение
                                try:
                                    numeric_value = float(replacement['replace_value']) if replacement['replace_value'] != '' else 0.0
                                    cell.value = numeric_value
                                    # Устанавливаем формат ячейки как General (стандартный формат Excel)
                                    cell.number_format = 'General'  # Используем General вместо 0.00
                                    cell.data_type = 'n'  # Явно указываем тип данных как число
                                    logger.info(f"Принудительно заменено 'חסרים משקלים' на {numeric_value} (число) в поле веса")
                                    replacements_made += 1
                                    break
                                except (ValueError, TypeError):
                                    logger.warning(f"Не удалось преобразовать значение веса '{replacement['replace_value']}' в число")
                                    continue                            # Сохраняем значение с правильным типом данных
                            if replacement.get('is_formula', False):
                                # Для формулы TODAY() вычисляем текущую дату и устанавливаем как значение
                                from datetime import datetime
                                current_date = datetime.now().date()  # Получаем текущую дату как объект date
                                cell.value = current_date
                                # Устанавливаем формат ячейки для правильного отображения даты
                                cell.number_format = 'DD.MM.YYYY'  # Формат даты для Excel
                                logger.info(f"Установлена текущая дата в столбце '{target_col}' (колонка {col_idx}, строка {data_row_idx}): {current_date.strftime('%d.%m.%Y')}")
                            elif replacement['is_numeric'] and (replacement['replace_value'] != '' or replacement.get('force_replace', False)):
                                try:
                                    # Преобразуем в число (для force_replace используем 0 если значение пустое)
                                    if replacement['replace_value'] != '':
                                        numeric_value = float(replacement['replace_value'])
                                    else:
                                        numeric_value = 0.0

                                    # Явно приводим к правильному типу для конкретных полей
                                    if target_col == 'מוצרים מוכנים לאכילה':
                                        numeric_value = float(numeric_value)  # Вес всегда float
                                    elif target_col == 'סה"כ קרטונים':
                                        # Для количества оставляем как float, без округления
                                        numeric_value = float(numeric_value)  # Количество может быть дробным
                                    else:
                                        numeric_value = float(numeric_value)  # Остальные поля как float

                                    # Специальная обработка для поля веса - проверяем, есть ли в ячейке текст "חסרים משקלים"
                                    if target_col == 'מוצרים מוכנים לאכילה':
                                        # Очищаем ячейку полностью перед установкой нового значения
                                        cell.value = None
                                        cell.style = 'Normal'

                                        # Устанавливаем числовое значение
                                        cell.value = float(numeric_value)
                                        cell.data_type = 'n'
                                        cell.number_format = 'General'

                                        # Применяем правильный шрифт как в оригинальном файле
                                        from openpyxl.styles import Font
                                        cell.font = Font(name='Arial', size=9)

                                        # Убеждаемся, что значение сохранено правильно
                                        logger.info(f"Установлено значение веса: {cell.value} (тип: {type(cell.value)})")

                                        # Дополнительная проверка - если значение все еще не число, конвертируем
                                        if not isinstance(cell.value, (int, float)):
                                            try:
                                                cell.value = float(numeric_value)
                                                logger.info(f"Принудительно конвертировано в число: {cell.value}")
                                            except Exception:
                                                logger.error("Не удалось установить числовое значение для веса")

                                        logger.info(f"Заменено поле веса с '{old_value}' на {numeric_value} (число) - специальный текст")
                                    # Специальная обработка для поля 'סה"כ משקל'
                                    elif target_col == 'סה"כ משקל':
                                        # Очищаем ячейку полностью перед установкой нового значения
                                        cell.value = None
                                        cell.style = 'Normal'

                                        # Устанавливаем числовое значение
                                        cell.value = float(numeric_value)
                                        cell.data_type = 'n'
                                        cell.number_format = 'General'

                                        # Применяем правильный шрифт как в оригинальном файле
                                        from openpyxl.styles import Font
                                        cell.font = Font(name='Arial', size=9)
                                        logger.info(f"Установлено значение общего веса: {cell.value} (тип: {type(cell.value)})")
                                        logger.info(f"Заменено поле общего веса с '{old_value}' на {numeric_value} (число)")
                                    # Специальная обработка для поля количества (аризות)
                                    elif target_col == 'סה"כ קרטונים':
                                        # Очищаем ячейку полностью перед установкой нового значения
                                        cell.value = None
                                        cell.style = 'Normal'

                                        # Устанавливаем числовое значение без округления
                                        cell.value = float(numeric_value)  # Оставляем как float, как в оригинале
                                        cell.data_type = 'n'
                                        cell.number_format = 'General'

                                        # Применяем правильный шрифт как в оригинальном файле
                                        from openpyxl.styles import Font
                                        cell.font = Font(name='Arial', size=9)

                                        # Убеждаемся, что значение сохранено правильно
                                        logger.info(f"Установлено значение количества: {cell.value} (тип: {type(cell.value)})")

                                        # Дополнительная проверка - если значение все еще не число, конвертируем
                                        if not isinstance(cell.value, (int, float)):
                                            try:
                                                cell.value = float(numeric_value)  # Оставляем как float
                                                logger.info(f"Принудительно конвертировано значение количества: {cell.value}")
                                            except Exception:
                                                logger.error("Не удалось установить значение для количества")

                                        logger.info(f"Заменено поле количества с '{old_value}' на {numeric_value} (дробное число)")
                                    elif numeric_value > 0:
                                        # Для других числовых полей устанавливаем только положительные значения
                                        cell.value = numeric_value
                                        # Устанавливаем числовой формат для всех числовых полей
                                        if target_col in ['סה"כ קרטונים']:
                                            cell.number_format = 'General'  # Используем General вместо 0.0
                                            cell.data_type = 'n'  # Явно указываем тип данных как число
                                            cell.style = 'Normal'  # Сбрасываем стиль
                                            # Применяем Arial шрифт для числовых полей
                                            from openpyxl.styles import Font
                                            cell.font = Font(name='Arial', size=9)
                                        elif target_col in ['מוצרים מוכנים לאכילה']:
                                            cell.number_format = 'General'  # Используем General вместо 0.0
                                            cell.data_type = 'n'  # Явно указываем тип данных как число
                                            cell.style = 'Normal'  # Сбрасываем стиль
                                            # Применяем Arial шрифт для числовых полей
                                            from openpyxl.styles import Font
                                            cell.font = Font(name='Arial', size=9)
                                        else:
                                            cell.number_format = '0'  # Целый формат для других числовых полей
                                            cell.data_type = 'n'  # Явно указываем тип данных как число
                                            cell.style = 'Normal'  # Сбрасываем стиль
                                            # Применяем Arial шрифт для числовых полей
                                            from openpyxl.styles import Font
                                            cell.font = Font(name='Arial', size=9)
                                        logger.info(f"Заменено в столбце '{target_col}' (колонка {col_idx}, строка {data_row_idx}): '{old_value}' -> {numeric_value} (число)")
                                    else:
                                        logger.info(f"Пропущено нулевое значение в столбце '{target_col}' - оставлено: '{old_value}'")

                                except (ValueError, TypeError) as e:
                                    logger.warning(f"Не удалось преобразовать '{replacement['replace_value']}' в число: {e}")
                                    cell.value = replacement['replace_value']
                                    logger.info(f"Заменено в столбце '{target_col}' (колонка {col_idx}, строка {data_row_idx}): '{old_value}' -> '{replacement['replace_value']}' (текст)")
                            else:
                                # Сохраняем как текст
                                cell.value = replacement['replace_value']
                                logger.info(f"Заменено в столбце '{target_col}' (колонка {col_idx}, строка {data_row_idx}): '{old_value}' -> '{replacement['replace_value']}' (текст)")

                            replacements_made += 1
                            break
                    except Exception as e:
                        logger.debug(f"Ошибка при замене в строке {data_row_idx}: {e}")
                        continue

        logger.info(f"Всего сделано замен: {replacements_made}")
        return replacements_made

    def _add_missing_fields(self, worksheet, field_replacements: List[Dict], replacement_data: Dict) -> int:
        """
        Add missing fields to the worksheet that are required by the robot system.

        Args:
            worksheet: openpyxl worksheet object
            field_replacements: List of field replacement configurations
            replacement_data: Dictionary with replacement values

        Returns:
            Number of fields added
        """
        fields_added = 0

        # Найдем последнюю заполненную строку
        last_row = worksheet.max_row
        for row_idx in range(last_row, 0, -1):
            if any(cell.value is not None for cell in worksheet[row_idx]):
                last_row = row_idx
                break

        # Добавляем поле веса, если его нет
        weight_field_exists = False
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and 'מוצרים מוכנים לאכילה' in str(cell.value):
                    weight_field_exists = True
                    break
            if weight_field_exists:
                break

        if not weight_field_exists:
            # Добавляем поле веса в новую строку
            new_row_idx = last_row + 2  # Пропускаем одну строку для читаемости

            # Находим подходящее место для поля (колонка 1)
            weight_cell = worksheet.cell(row=new_row_idx, column=1)
            weight_cell.value = 'מוצרים מוכנים לאכילה'

            # Добавляем значение веса в следующую колонку
            value_cell = worksheet.cell(row=new_row_idx, column=2)
            weight_value = self._get_weight_value(replacement_data)
            if weight_value > 0:
                value_cell.value = float(weight_value)
                value_cell.number_format = 'General'  # Используем General вместо 0.0
                value_cell.data_type = 'n'
                # Применяем Arial шрифт
                from openpyxl.styles import Font
                value_cell.font = Font(name='Arial', size=9)
                logger.info(f"Добавлено поле веса 'מוצרים מוכנים לאכילה' со значением {weight_value} в строку {new_row_idx}")
                fields_added += 1

        # Добавляем поле количества, если его нет
        quantity_field_exists = False
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and 'סה"כ קרטונים' in str(cell.value):
                    quantity_field_exists = True
                    break
            if quantity_field_exists:
                break

        if not quantity_field_exists:
            # Добавляем поле количества в следующую строку
            new_row_idx = last_row + 3

            # Находим подходящее место для поля (колонка 1)
            quantity_cell = worksheet.cell(row=new_row_idx, column=1)
            quantity_cell.value = 'סה"כ קרטונים'

            # Добавляем значение количества в следующую колонку
            value_cell = worksheet.cell(row=new_row_idx, column=2)
            quantity_value = replacement_data.get('סה\'כ אריזות', replacement_data.get('סהכ אריזות', 0))
            if quantity_value > 0:
                value_cell.value = float(quantity_value)
                value_cell.number_format = 'General'  # Используем General вместо 0.0
                value_cell.data_type = 'n'
                # Применяем Arial шрифт
                from openpyxl.styles import Font
                value_cell.font = Font(name='Arial', size=9)
                logger.info(f"Добавлено поле количества 'סה\"כ קרטונים' со значением {quantity_value} в строку {new_row_idx}")
                fields_added += 1

        logger.info(f"Добавлено {fields_added} отсутствующих полей")
        return fields_added

    def _search_fields_in_all_cells(self, worksheet, field_replacements: List[Dict], replacement_data: Dict) -> int:
        """
        Search for fields in all cells of the worksheet, not just first two columns.

        Args:
            worksheet: openpyxl worksheet object
            field_replacements: List of field replacement configurations
            replacement_data: Dictionary with replacement values

        Returns:
            Number of replacements made
        """
        replacements_made = 0

        # First, identify which fields actually exist in the template
        existing_fields = set()
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value is not None:
                    cell_value = str(cell.value)
                    # Ищем как точное совпадение, так и частичное
                    for replacement in field_replacements:
                        target_col = replacement['target_column']
                        intermediate_field = replacement['intermediate_field']

                        # Проверяем целевое поле (для робота)
                        if target_col in cell_value:
                            existing_fields.add(target_col)

                        # Проверяем промежуточное поле (из данных)
                        if intermediate_field in cell_value:
                            existing_fields.add(intermediate_field)

                        # Ищем похожие поля для веса
                        if 'משקל' in cell_value or 'weight' in cell_value.lower():
                            existing_fields.add('weight_related')

        # Log which fields exist and which are missing
        all_search_fields = set()
        for replacement in field_replacements:
            all_search_fields.update(replacement['search_fields'])

        missing_fields = all_search_fields - existing_fields
        if missing_fields:
            logger.info(f"Пропускаем поиск следующих полей (отсутствуют в шаблоне): {list(missing_fields)}")

        if not existing_fields:
            logger.info("В шаблоне не найдено ни одного из искомых полей. Поиск пропущен.")
            return 0

        logger.info(f"Выполняем поиск только для существующих полей: {list(existing_fields)}")

        # Search only for existing fields
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            row_cells = list(row)

            # Ищем в каждой строке все поля, которые нужно заменить
            for replacement in field_replacements:
                # Проверяем, есть ли данные для замены
                if not replacement['replace_value'] or str(replacement['replace_value']).strip() == '':
                    continue

                for search_field in replacement['search_fields']:
                    # Ищем поле в текущей строке
                    field_found_in_row = False
                    field_col_idx = None

                    for col_idx, cell in enumerate(row_cells, 1):
                        if cell.value is not None:
                            cell_value = str(cell.value)
                            # Handle Hebrew special characters
                            normalized_cell = cell_value.replace('"', "'").replace('״', "'")
                            normalized_search = search_field.replace('"', "'").replace('״', "'")

                            # Ищем подстроку в тексте ячейки
                            if search_field in cell_value or normalized_search in normalized_cell:
                                field_found_in_row = True
                                field_col_idx = col_idx
                                logger.info(f"Найдено поле '{search_field}' в строке {row_idx}, колонке {col_idx}: '{cell_value}'")
                                break

                    # Если нашли поле в строке, ищем где заменить значение
                    if field_found_in_row:
                        # Стратегия 1: Заменяем в следующей колонке той же строки
                        if field_col_idx < len(row_cells):
                            value_cell = row_cells[field_col_idx]  # Следующая ячейка в той же строке
                            if value_cell.value is not None:
                                old_value = value_cell.value
                                # Специальная обработка для числовых полей
                                if replacement['is_numeric']:
                                    try:
                                        numeric_value = float(replacement['replace_value']) if replacement['replace_value'] != '' else 0.0
                                        value_cell.value = numeric_value
                                        value_cell.number_format = 'General'  # Используем General вместо 0.0
                                        value_cell.data_type = 'n'
                                        # Применяем Arial шрифт
                                        from openpyxl.styles import Font
                                        value_cell.font = Font(name='Arial', size=9)
                                        logger.info(f"Заменено числовое поле '{search_field}' с '{old_value}' на {numeric_value} (число)")
                                    except (ValueError, TypeError):
                                        value_cell.value = replacement['replace_value']
                                        logger.info(f"Заменено поле '{search_field}' с '{old_value}' на '{replacement['replace_value']}' (текст)")
                                else:
                                    value_cell.value = replacement['replace_value']
                                    logger.info(f"Заменено поле '{search_field}' с '{old_value}' на '{replacement['replace_value']}' (строка {row_idx}, колонка {field_col_idx + 1})")
                                replacements_made += 1
                                break

                        # Стратегия 2: Ищем пустую ячейку в той же строке для замены
                        for col_idx in range(len(row_cells)):
                            if col_idx + 1 != field_col_idx:  # Пропускаем колонку с названием поля
                                check_cell = row_cells[col_idx]
                                if check_cell.value is None or str(check_cell.value).strip() == '':
                                    check_cell.value = replacement['replace_value']
                                    replacements_made += 1
                                    logger.info(f"Заменено поле '{search_field}' на '{replacement['replace_value']}' (строка {row_idx}, колонка {col_idx + 1})")
                                    break
                        else:
                            # Стратегия 3: Заменяем в колонке с наибольшим номером в строке
                            last_cell = row_cells[-1]
                            if last_cell.value is not None:
                                old_value = last_cell.value
                                last_cell.value = replacement['replace_value']
                                replacements_made += 1
                                logger.info(f"Заменено поле '{search_field}' с '{old_value}' на '{replacement['replace_value']}' (строка {row_idx}, последняя колонка)")
                            break

        return replacements_made

    def validate_reports_structure(self, reports_dir: Path) -> Dict[str, List[str]]:
        """
        Validate the structure of report files.

        Args:
            reports_dir: Directory containing report files

        Returns:
            Dictionary with validation results
        """
        results = {
            'valid_files': [],
            'invalid_files': [],
            'missing_files': [],
            'errors': []
        }

        if not reports_dir.exists():
            results['errors'].append(f"Reports directory does not exist: {reports_dir}")
            return results

        # Find Excel files
        excel_files = list(reports_dir.glob("*.xlsx"))

        if not excel_files:
            results['errors'].append("No Excel files found in reports directory")
            return results

        for file_path in excel_files:
            try:
                # Try to load workbook
                workbook = openpyxl.load_workbook(file_path, data_only=False)

                # Basic validation
                if len(workbook.sheetnames) > 0:
                    results['valid_files'].append(file_path.name)
                else:
                    results['invalid_files'].append(file_path.name)

                workbook.close()

            except Exception as e:
                results['invalid_files'].append(file_path.name)
                results['errors'].append(f"Error validating {file_path.name}: {e}")

        return results

    def get_processing_summary(self, results: Dict[str, str]) -> Dict[str, Union[int, List[str]]]:
        """
        Generate summary of processing results.

        Args:
            results: Dictionary with processing results

        Returns:
            Summary dictionary
        """
        summary = {
            'total_processed': len(results),
            'successful_files': list(results.keys()),
            'output_files': list(results.values())
        }

        return summary

    # --- Compatibility methods for workflow ---
    def search_reports_by_content(self, business_licenses: List[str]) -> Dict[str, str]:
        """Search for report files by scanning cell values for license numbers."""
        found: Dict[str, str] = {}
        try:
            licenses_set = set(str(lic) for lic in business_licenses)
            for pattern in ("*.xlsx", "*.xlsm"):
                for file_path in self.base_dir.rglob(pattern):
                    try:
                        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        for ws in wb.worksheets:
                            for row in ws.iter_rows(values_only=True):
                                for cell in row:
                                    if cell is None:
                                        continue
                                    cell_str = str(cell)
                                    for lic in list(licenses_set - set(found.keys())):
                                        if lic in cell_str:
                                            found[lic] = str(file_path)
                                            break
                                if len(found) == len(licenses_set):
                                    break
                            if len(found) == len(licenses_set):
                                break
                        wb.close()
                    except Exception:
                        continue
                    if len(found) == len(licenses_set):
                        break
        except Exception as e:
            logger.error(f"Error searching reports by content: {e}")
        return found

    def search_reports_by_license(self, business_licenses: List[str]) -> Dict[str, str]:
        """Search for report files by license numbers in filenames."""
        found: Dict[str, str] = {}
        try:
            for lic in business_licenses:
                lic_str = str(lic)
                for pattern in ("*.xlsx", "*.xlsm"):
                    for file_path in self.base_dir.rglob(pattern):
                        if lic_str in file_path.name:
                            found[lic_str] = str(file_path)
                            break
                    if lic_str in found:
                        break
        except Exception as e:
            logger.error(f"Error searching reports by license: {e}")
        return found

    def copy_reports_to_output(self, intermediate_file: Optional[Path] = None, found_reports: Optional[Dict[str, str]] = None) -> bool:
        """Copy found reports to output directory with optional data replacement."""
        try:
            if intermediate_file and intermediate_file.exists() and found_reports:
                # Use found reports from search
                self._copy_found_reports_with_replacement(found_reports, intermediate_file)
                return True
            elif intermediate_file and intermediate_file.exists():
                self.process_reports(self.base_dir, intermediate_file)
                return True
            else:
                # Simple copy without replacement
                for pattern in ("*.xlsx", "*.xlsm"):
                    for file_path in self.base_dir.rglob(pattern):
                        try:
                            dest = self.output_dir / file_path.name
                            shutil.copy2(file_path, dest)
                        except Exception as e:
                            logger.error(f"Error copying {file_path}: {e}")
                return True
        except Exception as e:
            logger.error(f"Error in copy_reports_to_output: {e}")
            return False

    def log_detailed_statistics(self):
        """Log detailed statistics about processing."""
        logger.info("Detailed statistics logging completed")

    def get_copy_summary(self) -> Dict[str, Union[int, float]]:
        """Return summary of copy operations."""
        return {
            'total_files': self.copied_files_count,
            'avg_files_per_license': 0,  # Could be calculated if we track per-license counts
            'min_files_per_license': 0,
            'max_files_per_license': 0
        }

    def _copy_found_reports_with_replacement(self, found_reports: Dict[str, str], intermediate_file: Path):
        """Copy only found reports with data replacement."""
        try:
            # Load replacement data
            logger.info(f"Loading replacement data from: {intermediate_file}")
            license_data_map = self._load_intermediate_data(intermediate_file)

            if not license_data_map:
                logger.error("No valid replacement data found")
                return

            logger.info(f"Loaded data for {len(license_data_map)} unique license-address combinations")

            results = {}
            for license_num, file_path_str in found_reports.items():
                try:
                    file_path = Path(file_path_str)

                    # Find all data entries for this license
                    license_entries = [(k, v) for k, v in license_data_map.items() if v['license_num'] == license_num]

                    if not license_entries:
                        logger.warning(f"No replacement data found for license: {license_num}")
                        continue

                    # Sort by address for consistent ordering
                    license_entries.sort(key=lambda x: x[1].get('address_key', ''))

                    # Process each address combination
                    for idx, (unique_key, replacement_data) in enumerate(license_entries):
                        address_key = replacement_data.get('address_key', 'no_address')

                        # Create unique output filename
                        base_name = file_path.stem
                        if len(license_entries) > 1:
                            # Multiple addresses - include address in filename
                            if address_key != 'no_address':
                                new_name = f"{base_name}_{address_key}.xlsx"
                            else:
                                new_name = f"{base_name}_{idx+1}.xlsx"
                        else:
                            # Single address - keep original name
                            new_name = f"{base_name}.xlsx"

                        dest_path = self.output_dir / new_name

                        # Perform replacement
                        logger.info(f"Processing: {file_path.name} -> {new_name} (license: {license_num}, address: {address_key})")
                        success = self._copy_file_with_replacement(file_path, dest_path, replacement_data)

                        if success:
                            results[file_path.name] = str(dest_path)
                            logger.info(f"Successfully processed: {new_name}")
                            self.copied_files_count += 1
                        else:
                            logger.error(f"Failed to process: {new_name}")

                except Exception as e:
                    logger.error(f"Error processing {file_path_str}: {e}")

            logger.info(f"Processing complete. Successfully processed {len(results)} files")

            # Report unprocessed licenses if any
            self._report_unprocessed_licenses_new(license_data_map, results)

        except Exception as e:
            logger.error(f"Error in _copy_found_reports_with_replacement: {e}")
